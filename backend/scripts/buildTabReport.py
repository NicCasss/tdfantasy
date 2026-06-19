# -*- coding: utf-8 -*-
"""Ricostruisce TabReport con un tab autonomo per ogni giornata.

Obiettivi:
- ogni giornata e' un foglio "Giornata N" indipendente (cambiare una giornata
  non tocca le altre);
- il compilatore inserisce solo PLAYER_ID, EVENTO, QTA (+ NOTE opzionale);
- tutto il resto si calcola da solo con ARRAYFORMULA (righe illimitate, niente
  formule da trascinare);
- menu a tendina per EVENTO e SQUADRE per ridurre gli errori di battitura.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

SRC = r"C:/Users/nicol/Downloads/TabReportNew (1).xlsx"
OUT = r"C:/Users/nicol/Downloads/TabReport_Migliorato.xlsx"

NUM_GIORNATE = 6  # = tournamentDays di default

src = openpyxl.load_workbook(SRC, data_only=False)

out = openpyxl.Workbook()
out.remove(out.active)

# ---- stili ----
H_INPUT = PatternFill("solid", fgColor="F26A00")   # header colonne da compilare
H_AUTO = PatternFill("solid", fgColor="9CA3AF")     # header colonne automatiche
C_INPUT = PatternFill("solid", fgColor="FFF3E0")    # celle da compilare
C_AUTO = PatternFill("solid", fgColor="F3F4F6")     # celle automatiche
LABEL = PatternFill("solid", fgColor="FDE8D5")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")


def copy_sheet(name):
    ws_src = src[name]
    ws = out.create_sheet(name)
    for row in ws_src.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            ws.cell(row=cell.row, column=cell.column, value=cell.value)
    for r in range(1, ws_src.max_column + 1):
        ws.cell(row=1, column=r).font = BOLD
    return ws


# ---- fogli sorgente ----
copy_sheet("PLAYERS")
copy_sheet("BONUS_MALUS")

# distinte squadre per il menu a tendina
teams = []
pl = src["PLAYERS"]
for r in range(2, pl.max_row + 1):
    v = pl.cell(row=r, column=3).value
    if v and v not in teams:
        teams.append(v)

# foglio liste (per i menu a tendina)
liste = out.create_sheet("LISTE")
liste["A1"] = "SQUADRE"
liste["A1"].font = BOLD
for i, t in enumerate(teams):
    liste.cell(row=2 + i, column=1, value=t)
liste.sheet_state = "hidden"

# range con nome usati dalle convalide
out.defined_names.add(
    DefinedName("EVENTI", attr_text="BONUS_MALUS!$B$2:$B$200")
)
out.defined_names.add(
    DefinedName("SQUADRE", attr_text=f"LISTE!$A$2:$A${1 + len(teams)}")
)


def build_giornata(n):
    ws = out.create_sheet(f"Giornata {n}")

    # ---- intestazione della giornata (valori, non formule globali) ----
    head = [
        ("GIORNATA", n),
        ("MODALITA", "PARTITA"),
        ("SQUADRA_A", ""),
        ("SQUADRA_B", ""),
    ]
    for i, (k, v) in enumerate(head):
        r = i + 1
        ws.cell(row=r, column=1, value=k).font = BOLD
        ws.cell(row=r, column=1).fill = LABEL
        c = ws.cell(row=r, column=2, value=v)
        c.fill = C_INPUT
        c.border = BORDER

    ws["D1"] = 'Compila SOLO le colonne arancioni: PLAYER_ID, EVENTO, QTA.'
    ws["D2"] = 'MODALITA: "PARTITA" filtra per le 2 squadre, "TUTTI" mostra tutti.'
    ws["D3"] = 'A destra (col. M-O) la lista degli atleti disponibili.'
    for rr in (1, 2, 3):
        ws.cell(row=rr, column=4).font = Font(italic=True, color="6B7280")

    # menu a tendina per SQUADRA_A / SQUADRA_B
    dv_sq = DataValidation(type="list", formula1="=SQUADRE", allow_blank=True)
    ws.add_data_validation(dv_sq)
    dv_sq.add(ws["B3"])
    dv_sq.add(ws["B4"])
    dv_mod = DataValidation(
        type="list", formula1='"PARTITA,TUTTI"', allow_blank=False
    )
    ws.add_data_validation(dv_mod)
    dv_mod.add(ws["B2"])

    # ---- tabella eventi (riga header = 6, dati da 7) ----
    HROW = 6
    headers = [
        ("PLAYER_ID", "input"),
        ("EVENTO", "input"),
        ("QTA", "input"),
        ("NOME", "auto"),
        ("SQUADRA", "auto"),
        ("VALORE_UNITARIO", "auto"),
        ("TOTALE", "auto"),
        ("GIORNATA", "auto"),
        ("SQUADRA_A", "auto"),
        ("SQUADRA_B", "auto"),
        ("NOTE", "input"),
    ]
    for i, (name, kind) in enumerate(headers):
        c = ws.cell(row=HROW, column=1 + i, value=name)
        c.font = WHITE
        c.alignment = CENTER
        c.border = BORDER
        c.fill = H_INPUT if kind == "input" else H_AUTO

    # ARRAYFORMULA: una sola formula riempie tutta la colonna
    F = HROW + 1  # 7
    af = {
        # NOME (col D) da PLAYERS
        4: f'=ARRAYFORMULA(IF(A{F}:A="","",IFERROR(VLOOKUP(A{F}:A,PLAYERS!$A:$C,2,FALSE),"PLAYER_ID non trovato")))',
        # SQUADRA (col E) da PLAYERS
        5: f'=ARRAYFORMULA(IF(A{F}:A="","",IFERROR(VLOOKUP(A{F}:A,PLAYERS!$A:$C,3,FALSE),"")))',
        # VALORE_UNITARIO (col F) da BONUS_MALUS (descrizione -> valore)
        6: f'=ARRAYFORMULA(IF(B{F}:B="","",IFERROR(VLOOKUP(B{F}:B,BONUS_MALUS!$B:$C,2,FALSE),"EVENTO non trovato")))',
        # TOTALE (col G) = QTA * VALORE_UNITARIO
        7: f'=ARRAYFORMULA(IF(B{F}:B="","",IFERROR(C{F}:C*F{F}:F,0)))',
        # GIORNATA (col H) dalla testa del foglio
        8: f'=ARRAYFORMULA(IF(A{F}:A="","",$B$1))',
        # SQUADRA_A (col I)
        9: f'=ARRAYFORMULA(IF(A{F}:A="","",$B$3))',
        # SQUADRA_B (col J)
        10: f'=ARRAYFORMULA(IF(A{F}:A="","",$B$4))',
    }
    for col, formula in af.items():
        ws.cell(row=F, column=col, value=formula)

    # lista atleti disponibili (riferimento per chi compila) in M:O
    ws.cell(row=HROW, column=13, value="ATLETI DISPONIBILI").font = BOLD
    ws.cell(
        row=F,
        column=13,
        value=(
            '=IFERROR(IF($B$2="TUTTI",'
            "FILTER(PLAYERS!A2:C,PLAYERS!E2:E=TRUE),"
            "FILTER(PLAYERS!A2:C,((PLAYERS!C2:C=$B$3)+(PLAYERS!C2:C=$B$4))*(PLAYERS!E2:E=TRUE))),"
            '"Imposta SQUADRA_A / SQUADRA_B")'
        ),
    )

    # menu a tendina EVENTO sulla colonna B (eventi)
    dv_ev = DataValidation(type="list", formula1="=EVENTI", allow_blank=True)
    ws.add_data_validation(dv_ev)
    dv_ev.add(f"B{F}:B1000")

    # estetica colonne dati: evidenzia input vs auto su molte righe
    for rr in range(F, 1001):
        for i, (_, kind) in enumerate(headers):
            cell = ws.cell(row=rr, column=1 + i)
            cell.fill = C_INPUT if kind == "input" else C_AUTO

    # larghezze e blocco riquadri
    widths = {
        "A": 12, "B": 22, "C": 6, "D": 22, "E": 16, "F": 16,
        "G": 10, "H": 10, "I": 14, "J": 14, "K": 18, "L": 3,
        "M": 12, "N": 22, "O": 14,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A7"
    return ws


# ---- foglio istruzioni ----
istr = out.create_sheet("ISTRUZIONI", 0)
righe = [
    ("FANTACORRADO - COMPILAZIONE REFERTI", True),
    ("", False),
    ("Per ogni giornata usa il foglio 'Giornata N' corrispondente.", False),
    ("Ogni giornata e' indipendente: quello che scrivi non tocca le altre.", False),
    ("", False),
    ("In testa al foglio imposta:", True),
    ("  - SQUADRA_A e SQUADRA_B (menu a tendina) della partita.", False),
    ("  - MODALITA: PARTITA (solo le 2 squadre) oppure TUTTI.", False),
    ("", False),
    ("Nella tabella compila SOLO le colonne arancioni:", True),
    ("  - PLAYER_ID  (vedi lista 'ATLETI DISPONIBILI' a destra)", False),
    ("  - EVENTO     (menu a tendina dai BONUS_MALUS)", False),
    ("  - QTA        (quante volte e' accaduto l'evento)", False),
    ("  - NOTE       (facoltativo)", False),
    ("", False),
    ("Le colonne grigie (NOME, SQUADRA, VALORE, TOTALE...) si compilano da sole.", False),
    ("Puoi aggiungere righe all'infinito: le formule si estendono in automatico.", False),
    ("", False),
    ("Quando l'evento o il player non vengono trovati appare un messaggio di", False),
    ("errore nella riga: correggi PLAYER_ID o EVENTO.", False),
]
for i, (txt, bold) in enumerate(righe):
    c = istr.cell(row=i + 1, column=1, value=txt)
    if bold:
        c.font = Font(bold=True, size=12, color="F26A00") if i == 0 else BOLD
istr.column_dimensions["A"].width = 80

for n in range(1, NUM_GIORNATE + 1):
    build_giornata(n)

out.save(OUT)
print("Salvato:", OUT)
print("Fogli:", out.sheetnames)
